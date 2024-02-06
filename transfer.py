import openpyxl
import os
import pinyin
import shutil


def read_ios_strings(file_path):
    print("read_ios_strings 开始读取{}文件".format(file_path))
    strings = {}
    with open(file_path, 'r') as file:
        lines = file.readlines()
        for line in lines:
            line = line.strip()
            if '=' in line:
                key, value = line.split('=', 1)
                strings[key.strip().strip('"')] = value.strip().strip(';').strip('"')
    print("read_ios_strings 一共读取 {} 行内容".format(len(strings)))
    return strings

def write_to_xlsx(file_path, keys, values):
    # print("write_to_xlsx 开始写{}行到{}文件中".format(len(keys),file_path))
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, key in enumerate(keys):
        sheet.cell(row=i+1, column=1, value=key)
        sheet.cell(row=i+1, column=2, value=values[i])
    workbook.save(file_path)

def write_to_strings_files(key, values, strings_files):
    for i, strings_file in enumerate(strings_files):
        with open(strings_file, 'a') as file:
            # print("write_to_strings_files 写的内容是 {} ++ {} 到文件{}中".format(key, values[i], strings_file))
            file.write('"%s" = "%s";\n' % (key, values[i]))

def find_value_in_xlsx(file_path, sheet_name, column_index, value):
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        cell_value = row[column_index]
        cell_pinyin = pinyin.get(cell_value, format="strip", delimiter="")
        value_pinyin = pinyin.get(value, format="strip", delimiter="")
        if cell_value == value:
            # print("find_value_in_xlsx 的{} 找到了{}, 翻译对应为{}".format(value, cell_value, row[3:6]))
            return row[2:5]
    return None

def process_ios_strings(file_path, xlsx_file_path, sheet_name, column_index, strings_files, redundant_xlsx_file):
    print("process_ios_strings 开始处理")
    strings = read_ios_strings(file_path)
    keys = []
    values = []
    for key, value in strings.items():
        result = find_value_in_xlsx(xlsx_file_path, sheet_name, column_index, value)
        if result is None:
            keys.append(key)
            values.append(value)
            continue
        else:
            write_to_strings_files(key, result, strings_files)
    write_to_xlsx(redundant_xlsx_file, keys, values)

def remove_file(folder_path):
    # 列出文件夹中的所有文件
    file_list = os.listdir(folder_path)
    # 逐个删除文件
    for file_name in file_list:
        file_path = os.path.join(folder_path, file_name)
        os.remove(file_path)

# 需要国际化文件所在的目录
ios_strings_file_folder_path = './source'
# 最终国际化文件存放的目录
ios_strings_file_final_path = './final'
# 需要国际化的文件
ios_strings_files_array = ['InfoPlist','InfoPlist_t','zh','LiveClass','DocCenter','Share','Homework','Localizable']
# 定义国际化翻译所在的 XLSX 文件路径
xlsx_file = '{}/国际化翻译.xlsx'.format(ios_strings_file_folder_path)
# 定义 XLSX 文件中要查找的列索引（从0开始）
column_index = 1
# 国际化翻译结果的临时存放目录
output_folder_path = './output'
# 定义要写入的.strings文件路径
strings_files = ['{}/ja.strings'.format(output_folder_path), '{}/ko.strings'.format(output_folder_path), '{}/de.strings'.format(output_folder_path)]
# 没有找到对应国际化翻译的字符串
redundant_xlsx_file = '{}/new_strings.xlsx'.format(output_folder_path)
sheet_name = '云学堂'

for i, strings_file_name in enumerate(ios_strings_files_array):
    print("开始处理{}文件".format(strings_file_name))
    remove_file(output_folder_path)
    # 处理 iOS 的国际化文件
    ios_strings_file = '{}/{}.strings'.format(ios_strings_file_folder_path,strings_file_name)
    process_ios_strings(ios_strings_file, xlsx_file, sheet_name, column_index, strings_files, redundant_xlsx_file)
    dest_file_path = '{}/{}/'.format(ios_strings_file_final_path,strings_file_name)
    remove_file(dest_file_path)
    shutil.move(redundant_xlsx_file, dest_file_path)
    for new_strings_file in strings_files:
        if os.path.exists(new_strings_file):
            print("移动到{}时，移动成功".format(dest_file_path))
            shutil.move(new_strings_file, dest_file_path)
        else:
            print("移动到{}时，{}文件不存在".format(dest_file_path,new_strings_file))


